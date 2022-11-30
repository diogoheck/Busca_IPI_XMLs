from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


NOME_PLANILHA = 'soma_IPI_notas.xlsx'


def autosize(ws):
    for i in range(1, ws.max_column+1):
        ws.column_dimensions[get_column_letter(i)].bestFit = True


def formata_numero(ws, intervalo):
    columns = ws[intervalo]

    for rows in columns:
        for cell in rows:
            cell.number_format = '#,##0.00'


def criar_pasta_trabalho():
    wb = Workbook()
    ws = wb.active
    ws.title = 'IPI_soma'
    wb.save(NOME_PLANILHA)
    return wb


def salvar_planilha_IPI(notas, lista_chaves_sped):
    wb = load_workbook(NOME_PLANILHA)

    ws = wb.active

    i = 1
    ws[f'A{i}'] = 'DATA'
    ws[f'B{i}'] = 'CHAVE'
    ws[f'C{i}'] = 'NOTA'
    ws[f'D{i}'] = 'CFOP'
    ws[f'E{i}'] = 'VLR IPI'

    for nf in notas.values():
        if nf[1][0:44] in lista_chaves_sped:
            ws.append(nf)

    # formata_numero(ws, 'E:E')
    # ws.merge_cells('A1:G1')
    ws.auto_filter.ref = 'A:E'
    autosize(ws)

    wb.save(NOME_PLANILHA)
